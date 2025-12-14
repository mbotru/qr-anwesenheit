from flask import Flask, request, render_template_string, send_file
from openpyxl import Workbook, load_workbook
from datetime import datetime
from geopy.geocoders import Nominatim
import qrcode
import os

app = Flask(__name__)

EXCEL_FILE = "anwesenheit.xlsx"
QR_FILE = "qr.png"

geolocator = Nominatim(user_agent="anwesenheit_app")

def get_city(lat, lon):
    try:
        location = geolocator.reverse(
            f"{lat}, {lon}",
            language="de",
            exactly_one=True
        )
        if not location:
            return "Unbekannt"

        addr = location.raw.get("address", {})
        return (
            addr.get("city")
            or addr.get("town")
            or addr.get("village")
            or addr.get("municipality")
            or "Unbekannt"
        )
    except:
        return "Unbekannt"


# Excel anlegen
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Datum/Uhrzeit", "Ort"])
    wb.save(EXCEL_FILE)


def create_qr(url):
    img = qrcode.make(url)
    img.save(QR_FILE)


@app.route("/")
def index():
    url = request.host_url + "checkin"
    create_qr(url)

    return render_template_string("""
    <h2>QR-Code scannen</h2>
    <img src="/qr" width="300">
    <p>{{ url }}</p>
    """, url=url)


@app.route("/qr")
def qr():
    return send_file(QR_FILE, mimetype="image/png")


@app.route("/checkin", methods=["GET", "POST"])
def checkin():
    if request.method == "POST":
        name = request.form.get("name")
        lat = request.form.get("lat")
        lon = request.form.get("lon")

        city = get_city(lat, lon)
        time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append([name, time, city])
        wb.save(EXCEL_FILE)

        return f"<h3>✅ Gespeichert – Ort: {city}</h3>"

    return render_template_string("""
    <h2>Anwesenheit</h2>
    <form method="POST">
        <input name="name" placeholder="Name" required><br><br>
        <input type="hidden" name="lat" id="lat">
        <input type="hidden" name="lon" id="lon">
        <button>Einchecken</button>
    </form>

    <script>
    navigator.geolocation.getCurrentPosition(
        pos => {
            lat.value = pos.coords.latitude;
            lon.value = pos.coords.longitude;
        },
        err => alert("Standortfreigabe erforderlich!")
    );
    </script>
    """)
    

if __name__ == "__main__":
    app.run()

